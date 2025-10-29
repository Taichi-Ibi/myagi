import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def analyze_issue(no, business, request, dept):
    """各課題を分析して技術・難易度・実現方法を返す"""

    # 分析ロジック
    tech = []
    difficulty = 1
    method = ""

    # 各課題の分析
    if no == 1:
        tech = ["デザインツール", "画像生成AI"]
        difficulty = 2
        method = "生成AIによるスライドマスターデザイン生成、デザインツールでの調整・著作権確保"
    elif no == 2:
        tech = ["生成AI", "Excel/Python"]
        difficulty = 3
        method = "ExcelマクロまたはPythonで試算ロジック構築、生成AIに試算条件とロジックを提示し繰り返し実行・精度向上"
    elif no == 3:
        tech = ["Webスクレイピング", "生成AI", "BIツール"]
        difficulty = 3
        method = "Webスクレイピングで市場情報収集、生成AIで構造化・分析、BIツールで可視化"
    elif no == 4:
        tech = ["生成AI"]
        difficulty = 2
        method = "製品情報・市場データを生成AIに提示し、戦略オプションと論理的根拠をアウトプット"
    elif no == 5:
        tech = ["API連携", "論文DB"]
        difficulty = 2
        method = "Google Scholar API、PubMed API等の論文DBと連携し検索インターフェース構築"
    elif no == 6:
        tech = ["生成AI"]
        difficulty = 1
        method = "資料をテキスト化し生成AIでレビュー、論理性・網羅性・説得力の観点でフィードバック生成"
    elif no == 7:
        tech = ["生成AI"]
        difficulty = 1
        method = "資料を生成AIに入力し、体裁・表現・可読性の観点でレビューコメント生成"
    elif no == 8:
        tech = ["Salesforce API", "BIツール", "Python"]
        difficulty = 2
        method = "Salesforce APIでデータ取得、Pythonで加工、Tableau/PowerBI等でグラフ・図を自動生成"
    elif no == 9:
        tech = ["Webスクレイピング", "生成AI"]
        difficulty = 1
        method = "Webスクレイピングで市場調査資料取得、生成AIで要点抽出・要約"
    elif no == 10:
        tech = ["Web検索API", "生成AI"]
        difficulty = 2
        method = "仮説をクエリ化しWeb検索API実行、生成AIで情報の信頼性評価・ソース提示"
    elif no == 11:
        tech = ["生成AI"]
        difficulty = 1
        method = "製品情報・ターゲット・訴求軸を生成AIに提示し、複数パターンのコピー案を大量生成"
    elif no == 12:
        tech = ["Webスクレイピング", "生成AI", "スケジューラ"]
        difficulty = 2
        method = "競合企業のWeb・SNSを定期スクレイピング、生成AIで発信内容分類・サマリ化、レポート自動生成"
    elif no == 13:
        tech = ["生成AI", "Web検索API"]
        difficulty = 2
        method = "業界情報をWeb検索API取得、生成AIで課題・ハードル・リスクをリスト化"
    elif no == 14:
        tech = ["生成AI"]
        difficulty = 1
        method = "ヒアリング目的・顧客属性を生成AIに提示し、質問リスト自動生成"
    elif no == 15:
        tech = ["生成AI", "文字起こしAPI"]
        difficulty = 1
        method = "議事メモまたは音声を文字起こし、生成AIで顧客課題・示唆・次アクション抽出"
    elif no == 16:
        tech = ["生成AI"]
        difficulty = 1
        method = "既存スライドテキストと訴求対象を生成AIに提示し、より刺さる表現にリライト"
    elif no == 17:
        tech = ["生成AI"]
        difficulty = 1
        method = "提案要件を生成AIに提示し、目的・課題・解決策・効果の骨子を自動生成"
    elif no == 18:
        tech = ["生成AI", "社内DB連携"]
        difficulty = 2
        method = "顧客情報・過去議事録をDBから取得、生成AIでカスタマイズした提案文を下書き"
    elif no == 19:
        tech = ["Web検索API", "生成AI", "GAS"]
        difficulty = 2
        method = "GASで毎朝定時実行、Web検索APIでニュース取得、生成AIでサマリ化・メール送信"
    elif no == 20:
        tech = ["生成AI", "テンプレート"]
        difficulty = 1
        method = "メール種別と必要情報を入力、生成AIがテンプレートベースでドラフト自動生成"
    elif no == 21:
        tech = ["生成AI"]
        difficulty = 1
        method = "元文案とトーン指定を生成AIに提示し、複数候補を生成"
    elif no == 22:
        tech = ["生成AI", "テキスト分類"]
        difficulty = 2
        method = "営業トーク・質問履歴を生成AIで分類・構造化、FAQ形式で整形・更新"
    elif no == 23:
        tech = ["生成AI"]
        difficulty = 1
        method = "ドキュメントを生成AIに入力し、重要情報を抽出して3分で読める要約生成"
    elif no == 24:
        tech = ["文字起こしAPI", "生成AI"]
        difficulty = 2
        method = "会議音声を文字起こしAPI処理、生成AIで議事録整形・次アクション抽出"
    elif no == 25:
        tech = ["ベクトル検索", "生成AI", "RAG"]
        difficulty = 3
        method = "過去資料をベクトル化しDB登録、RAGシステムで類似案件を横断検索・要約"
    elif no == 26:
        tech = ["生成AI", "DB"]
        difficulty = 2
        method = "社内文書から用語抽出、生成AIで略称・正式名称マッピング、用語辞書DB構築"
    elif no == 27:
        tech = ["生成AI"]
        difficulty = 1
        method = "専門用語資料とリテラシーレベルを生成AIに提示し、平易な表現に変換"
    elif no == 28:
        tech = ["生成AI"]
        difficulty = 1
        method = "新機能情報を生成AIに提示し、プレスリリース形式で下書き生成"
    elif no == 29:
        tech = ["翻訳API", "生成AI"]
        difficulty = 1
        method = "DeepL API等の翻訳APIで一次翻訳、生成AIで文脈・業界用語調整"
    elif no == 30:
        tech = ["Web検索API", "生成AI"]
        difficulty = 2
        method = "Web検索APIで市場情報収集、生成AIで規模・プレイヤー・特徴等を構造化し1枚サマリ生成"
    elif no == 31:
        tech = ["生成AI", "テキスト分類"]
        difficulty = 2
        method = "リード情報・問い合わせ内容を生成AIで自動分類、カテゴリタグ付与"
    elif no == 32:
        tech = ["生成AI", "スコアリング"]
        difficulty = 2
        method = "案件情報をスコアリングルールまたは生成AIで評価、優先度アラート自動送信"
    elif no == 33:
        tech = ["生成AI", "データ連携"]
        difficulty = 1
        method = "進捗データを収集、生成AIで定例会用スライドドラフト生成"
    elif no == 34:
        tech = ["生成AI"]
        difficulty = 1
        method = "FAQ素材を生成AIでユーザー視点の平易な表現に整形"
    elif no == 35:
        tech = ["生成AI", "RAG"]
        difficulty = 2
        method = "FAQ・製品情報をRAGシステムに登録、問い合わせメールから一次回答案自動生成"
    elif no == 36:
        tech = ["生成AI"]
        difficulty = 1
        method = "議事録を生成AIに入力し、目的・決定事項・TODO抽出"
    elif no == 37:
        tech = ["API連携", "生成AI"]
        difficulty = 2
        method = "Slack API等でTODO収集、生成AIで重複除去・優先順位付けし一覧化"
    elif no == 38:
        tech = ["生成AI", "カレンダーAPI"]
        difficulty = 2
        method = "候補日・議題・目的を入力、生成AIで日程調整メール文面生成、カレンダーAPI連携"
    elif no == 39:
        tech = ["生成AI", "データ連携"]
        difficulty = 1
        method = "進捗データ収集、生成AIで定型フォーマットに沿ってレポート自動生成"
    elif no == 40:
        tech = ["生成AI"]
        difficulty = 1
        method = "会議の目的・前提・ゴールを生成AIに提示し、アジェンダ叩き台生成"
    elif no == 41:
        tech = ["Webスクレイピング", "生成AI"]
        difficulty = 2
        method = "企業情報サイトをスクレイピング、生成AIで数値抽出・構造化し一覧化"
    elif no == 42:
        tech = ["生成AI"]
        difficulty = 1
        method = "長文資料を生成AIで要約・簡潔化し新人向け教育資料に変換"
    elif no == 43:
        tech = ["生成AI"]
        difficulty = 1
        method = "文書を生成AIでチェック、誤解・炎上リスクのある表現を指摘"
    elif no == 44:
        tech = ["生成AI", "Web検索API"]
        difficulty = 2
        method = "課題をクエリ化しWeb検索API実行、生成AIで背景・構造を整理し説明文生成"
    elif no == 45:
        tech = ["生成AI", "DB"]
        difficulty = 2
        method = "新規質問と既存FAQをベクトル化、生成AIで重複チェック・マージし最新FAQ維持"
    elif no == 46:
        tech = ["生成AI"]
        difficulty = 1
        method = "商談メモを生成AIに入力し、次回提案・優先課題・温度感を抽出"
    elif no == 47:
        tech = ["Salesforce API", "データ検証", "生成AI"]
        difficulty = 2
        method = "Salesforce APIでデータ取得、検証ルールまたは生成AIで異常値・欠損値検出"
    elif no == 48:
        tech = ["生成AI"]
        difficulty = 1
        method = "アナウンス内容を生成AIに提示し、適切な長さ・トーンの文案生成"
    elif no == 49:
        tech = ["生成AI", "Web検索API"]
        difficulty = 2
        method = "業界情報をWeb検索API取得、生成AIで刺さりポイントの初期仮説生成"
    elif no == 50:
        tech = ["生成AI"]
        difficulty = 2
        method = "サービスアイデアを生成AIに提示し、市場性・実現性・筋書きをブレスト形式で生成"
    elif no == 51:
        tech = ["メールAPI", "Salesforce API", "生成AI"]
        difficulty = 2
        method = "メールAPIで問い合わせ受信、生成AIで内容構造化、Salesforce APIで自動起票"
    elif no == 52:
        tech = ["生成AI", "CRM連携"]
        difficulty = 2
        method = "顧客属性をCRMから取得、生成AIで属性別フォロー内容提案"
    elif no == 53:
        tech = ["生成AI"]
        difficulty = 2
        method = "顧客セグメント情報を生成AIに提示し、セグメント別訴求メッセージ整理"
    elif no == 54:
        tech = ["生成AI"]
        difficulty = 1
        method = "価格・価値情報を生成AIに提示し、正当性説明の言い回し集生成"
    elif no == 55:
        tech = ["生成AI", "DB"]
        difficulty = 1
        method = "顧客反応メモを生成AIで構造化、刺さるワード抽出・ナレッジDB登録"
    elif no == 56:
        tech = ["生成AI"]
        difficulty = 1
        method = "イベント情報を生成AIに提示し、登壇資料・ブース資料・配布物ドラフト一括生成"
    elif no == 57:
        tech = ["生成AI", "CRM連携"]
        difficulty = 1
        method = "名刺情報・会話内容を生成AIに提示し、個別フォロー案自動生成"
    elif no == 58:
        tech = ["生成AI"]
        difficulty = 1
        method = "RFP・要件定義書を生成AIに入力し、真のニーズを抽出・要約"
    elif no == 59:
        tech = ["生成AI", "Excel/Python"]
        difficulty = 2
        method = "施策内容・前提条件を生成AIに提示し、効果試算ロジック・ラフ数値生成"
    elif no == 60:
        tech = ["生成AI", "OCR"]
        difficulty = 2
        method = "契約書をOCR処理、生成AIで記入項目チェック・漏れ検出"
    elif no == 61:
        tech = ["Excel/GAS", "BIツール", "生成AI"]
        difficulty = 2
        method = "数値入力でグラフ自動生成、生成AIで良否判定コメント付与"
    elif no == 62:
        tech = ["Salesforce API", "BIツール", "生成AI"]
        difficulty = 2
        method = "Salesforce APIで案件データ取得、BIツールで可視化、生成AIで次アクション提案"
    elif no == 63:
        tech = ["生成AI", "データ連携"]
        difficulty = 1
        method = "月次活動データ収集、生成AIで報告書ドラフト自動生成"
    elif no == 64:
        tech = ["生成AI", "CRM連携", "スケジューラ"]
        difficulty = 2
        method = "CRMで更新時期検知、生成AIで状況確認文面生成・自動送信"
    elif no == 65:
        tech = ["生成AI", "テンプレート"]
        difficulty = 2
        method = "契約条件を入力、生成AIがテンプレートベースで契約書初稿生成"
    elif no == 66:
        tech = ["生成AI"]
        difficulty = 1
        method = "稟議内容を生成AIに提示し、背景・目的・リスク・依頼事項を整理"
    elif no == 67:
        tech = ["生成AI"]
        difficulty = 1
        method = "顧客成果情報を生成AIに提示し、事例シート形式で整形"
    elif no == 68:
        tech = ["生成AI"]
        difficulty = 2
        method = "価格構造・値下げ要望を生成AIに提示し、交渉論点・譲歩ライン整理"
    elif no == 69:
        tech = ["生成AI", "データ連携"]
        difficulty = 1
        method = "顧客向けデータ収集、生成AIでサマリレポート初稿生成"
    elif no == 70:
        tech = ["Web検索API", "生成AI"]
        difficulty = 1
        method = "企業名で最新ニュース検索、生成AIで要約・商談準備資料生成"
    elif no == 71:
        tech = ["生成AI"]
        difficulty = 2
        method = "提案内容・顧客属性を生成AIに提示し、決裁者の懸念点・リスク事前洗い出し"
    elif no == 72:
        tech = ["生成AI"]
        difficulty = 1
        method = "提案資料を生成AIに入力し、論理の飛躍・一貫性欠如を検証"
    elif no == 73:
        tech = ["生成AI"]
        difficulty = 1
        method = "MTG目的・ゴールを生成AIに提示し、会議アジェンダ生成"
    elif no == 74:
        tech = ["Salesforce API", "生成AI", "Excel"]
        difficulty = 2
        method = "Salesforce APIで案件進捗取得、生成AIまたはExcelで売上予測算出"
    elif no == 75:
        tech = ["生成AI", "機械学習"]
        difficulty = 3
        method = "顧客データ（利用状況・契約情報等）を機械学習または生成AIでスコアリング"
    elif no == 76:
        tech = ["Salesforce API", "生成AI"]
        difficulty = 2
        method = "Salesforce・議事録からタスク抽出、生成AIで抜け漏れ検出・アラート"
    elif no == 77:
        tech = ["BIツール", "Excel/GAS"]
        difficulty = 1
        method = "目標・実績データをBIツールまたはExcelで可視化、達成率表示"
    elif no == 78:
        tech = ["Salesforce API", "生成AI"]
        difficulty = 2
        method = "商談後に必須項目チェック、Salesforce APIで入力抜け検出・アラート"
    elif no == 79:
        tech = ["Salesforce API", "生成AI", "GAS"]
        difficulty = 2
        method = "Salesforce APIで数値取得、生成AIまたはGASで報告資料自動ドラフト"
    elif no == 80:
        tech = ["Salesforce API", "スケジューラ", "生成AI"]
        difficulty = 2
        method = "〆切データを一覧化、スケジューラで期限チェック、抜け警告自動送信"
    elif no == 81:
        tech = ["生成AI", "Salesforce API"]
        difficulty = 1
        method = "商談メモ・議事録・メールを生成AIで構造化、Salesforce入力項目に変換・下書き"
    else:
        tech = ["生成AI"]
        difficulty = 1
        method = "生成AIで処理"

    return tech, difficulty, method

def create_excel():
    """JSONLファイルを読み込み、分析結果を追加してExcelを生成"""

    # JSONLファイル読み込み
    issues = []
    with open('issues.jsonl', 'r', encoding='utf-8') as f:
        for line in f:
            issues.append(json.loads(line))

    # Excelワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "マーケティング課題分析"

    # ヘッダー設定
    headers = [
        'No',
        '対象業務',
        '求めること',
        '部署',
        '想定される主な使用技術',
        '技術的難易度',
        '実現方法'
    ]

    # ヘッダー行のスタイル
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # データ追加
    for issue in issues:
        no = issue['No']
        business = issue['対象業務']
        request = issue['求めること']
        dept = issue['部署']

        # 分析実行
        tech, difficulty, method = analyze_issue(no, business, request, dept)
        tech_str = ", ".join(tech)

        row_data = [no, business, request, dept, tech_str, difficulty, method]
        row_idx = no + 1

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # 難易度は中央揃え
            if col_idx == 6:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # 列幅調整
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 60

    # 行の高さ調整（ヘッダー）
    ws.row_dimensions[1].height = 30

    # ファイル保存
    output_file = 'マーケティング課題分析結果.xlsx'
    wb.save(output_file)
    print(f"✓ Excelファイルを生成しました: {output_file}")
    print(f"✓ 分析対象: {len(issues)}件の課題")

if __name__ == '__main__':
    create_excel()
